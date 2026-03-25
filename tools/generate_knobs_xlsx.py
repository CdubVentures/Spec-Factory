"""Generate SPEC_FACTORY_KNOBS.xlsx from the live settings registry.

Runs Node.js to dump RUNTIME_SETTINGS_REGISTRY, BOOTSTRAP_ENV_REGISTRY,
UI_SETTINGS_REGISTRY, and STORAGE_SETTINGS_REGISTRY as JSON, then builds
a multi-tab Excel workbook grouped by uiSection / group.

Usage:
    python tools/generate_knobs_xlsx.py
"""

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = REPO_ROOT / "docs" / "implementation" / "ai-indexing-plans" / "pipeline" / "SPEC_FACTORY_KNOBS.xlsx"

# ── Styles ─────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10, color="1F3864")

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)

SECRET_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DEPRECATED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
BOOL_TRUE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BOOL_FALSE_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
READONLY_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
RETIRED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# ── uiSection -> human-readable tab name ───────────────────────────────

SECTION_TAB_NAMES = {
    "run-setup": "Run Setup",
    "output": "Output & Automation",
    "search-profile": "Search Profile",
    "search-execution": "Search Execution",
    "search-planner": "Search Planner",
    "serp-selector": "SERP Selector",
    "domain-classifier": "Domain Classifier",
    "needset": "NeedSet",
    "browser": "Browser & Rendering",
    "network": "Network & Fetch",
    "adapter": "Fetcher Adapter",
    "screenshots": "Screenshots",
    "observability": "Observability",
    "schema": "Schema Enforcement",
    "models": "LLM Models",
    "limits": "LLM Limits",
    "provider": "LLM Provider",
}

# Stable tab order for runtime sections
SECTION_ORDER = list(SECTION_TAB_NAMES.keys())

# Tab colors keyed by human-readable tab name
TAB_COLORS = {
    "Run Setup": "4472C4",
    "Output & Automation": "70AD47",
    "Search Profile": "2E75B6",
    "Search Execution": "5B9BD5",
    "Search Planner": "9DC3E6",
    "SERP Selector": "4BACC6",
    "Domain Classifier": "31859C",
    "NeedSet": "00B0F0",
    "Browser & Rendering": "5B9BD5",
    "Network & Fetch": "2E75B6",
    "Fetcher Adapter": "9DC3E6",
    "Screenshots": "ED7D31",
    "Observability": "ED7D31",
    "Schema Enforcement": "FFC000",
    "LLM Models": "7030A0",
    "LLM Limits": "9966FF",
    "LLM Provider": "BF8FFF",
    "Bootstrap": "A5A5A5",
    "Storage": "A5A5A5",
    "UI Settings": "00B0F0",
    "Retired": "C00000",
}

COLUMNS = ["Key", "Type", "Default", "Min", "Max", "Env Var", "Config Key", "Aliases", "Flags", "Description"]
COL_WIDTHS = [38, 10, 32, 12, 14, 40, 34, 24, 22, 52]

# ── Retired settings (removed 2026-03-24) ──────────────────────────────

RETIRED_SETTINGS = [
    ("awsRegion", "string", "us-east-2", "2026-03-24", "storage cleanup", "Moved to STORAGE_SETTINGS_REGISTRY"),
    ("s3Bucket", "string", "my-spec-harvester-data", "2026-03-24", "storage cleanup", "Moved to STORAGE_SETTINGS_REGISTRY"),
    ("mirrorToS3", "bool", "false", "2026-03-24", "storage cleanup", "Replaced by persistent storage system"),
    ("mirrorToS3Input", "bool", "false", "2026-03-24", "storage cleanup", "Replaced by persistent storage system"),
    ("outputMode", "enum", "local", "2026-03-24", "storage cleanup", "Replaced by persistent storage system"),
    ("s3InputPrefix", "string", "specs/inputs", "2026-03-24", "storage cleanup", "Replaced by persistent storage system"),
    ("s3OutputPrefix", "string", "specs/outputs", "2026-03-24", "storage cleanup", "Replaced by persistent storage system"),
    ("resumeMode", "enum", "auto", "2026-03-24", "resume purge", "Resume system removed"),
    ("resumeWindowHours", "int", "48", "2026-03-24", "resume purge", "Resume system removed"),
    ("indexingResumeSeedLimit", "int", "24", "2026-03-24", "resume purge", "Resume system removed"),
    ("indexingResumePersistLimit", "int", "160", "2026-03-24", "resume purge", "Resume system removed"),
    ("indexingResumeRetryPersistLimit", "int", "", "2026-03-24", "bootstrap resume", "Resume system removed"),
    ("indexingResumeSuccessPersistLimit", "int", "", "2026-03-24", "bootstrap resume", "Resume system removed"),
    ("localMode", "bool", "true", "2026-03-24", "output purge", "Replaced by persistent storage system"),
    ("categoryAuthorityEnabled", "bool", "true", "2026-03-24", "output purge", "Retired — category authority always active"),
    ("writeMarkdownSummary", "bool", "true", "2026-03-24", "output purge", "No longer used"),
    ("crawlBypassMinBodyLength", "int", "200", "2026-03-24", "dead bypass knobs", "Call sites never passed value — hardcoded fallback always used"),
    ("crawlBypassHtmlSnippetCap", "int", "5000", "2026-03-24", "dead bypass knobs", "Call sites never passed value — hardcoded fallback always used"),
]

# ── Node.js JSON dump ──────────────────────────────────────────────────

NODE_SCRIPT = """\
import { RUNTIME_SETTINGS_REGISTRY, BOOTSTRAP_ENV_REGISTRY, UI_SETTINGS_REGISTRY, STORAGE_SETTINGS_REGISTRY } from './src/shared/settingsRegistry.js';
const out = {
  runtime: RUNTIME_SETTINGS_REGISTRY,
  bootstrap: BOOTSTRAP_ENV_REGISTRY,
  ui: UI_SETTINGS_REGISTRY,
  storage: STORAGE_SETTINGS_REGISTRY,
};
console.log(JSON.stringify(out));
"""


def load_registry_json():
    """Run Node.js to dump the live registry as JSON."""
    result = subprocess.run(
        ["node", "--input-type=module", "-e", NODE_SCRIPT],
        capture_output=True,
        text=True,
        cwd=str(REPO_ROOT),
    )
    if result.returncode != 0:
        print(f"Node.js dump failed (exit {result.returncode}):", file=sys.stderr)
        print(result.stderr, file=sys.stderr)
        sys.exit(1)
    return json.loads(result.stdout)


# ── Entry -> row tuple conversion ──────────────────────────────────────

def _flags(entry):
    """Build a flags string from entry metadata."""
    parts = []
    if entry.get("uiHero"):
        parts.append("hero")
    if entry.get("secret"):
        parts.append("secret")
    if entry.get("readOnly"):
        parts.append("readOnly")
    if entry.get("defaultsOnly"):
        parts.append("defaultsOnly")
    if entry.get("computed"):
        parts.append("computed")
    if entry.get("tokenClamped"):
        parts.append("tokenClamped")
    pg = entry.get("policyGroup", "")
    if pg:
        parts.append(f"policyGroup: {pg}")
    return ", ".join(parts)


def _description(entry):
    """Build a description string from entry metadata."""
    parts = []
    if entry.get("uiHero"):
        parts.append("[HERO]")
    if entry.get("secret"):
        parts.append("[SECRET]")
    if entry.get("readOnly"):
        parts.append("[READ-ONLY]")
    if entry.get("defaultsOnly"):
        parts.append("[INTERNAL]")
    tip = entry.get("uiTip", "")
    if tip:
        parts.append(tip)
    disabled_by = entry.get("disabledBy", "")
    if disabled_by:
        parts.append(f"Disabled when {disabled_by}=false")
    aliases = entry.get("aliases", [])
    if aliases:
        parts.append(f"Aliases: {', '.join(aliases)}")
    return " ".join(parts)


def _default_display(val):
    """Format a default value for display."""
    if val is None:
        return "null"
    if isinstance(val, bool):
        return str(val).lower()
    if isinstance(val, str) and len(val) > 80:
        return val[:77] + "..."
    return str(val)


def runtime_entry_to_row(entry):
    """Convert a RUNTIME_SETTINGS_REGISTRY entry to a tuple row."""
    return (
        entry.get("key", ""),
        entry.get("type", ""),
        _default_display(entry.get("default", "")),
        entry.get("min", ""),
        entry.get("max", ""),
        entry.get("envKey", ""),
        entry.get("configKey", ""),
        ", ".join(entry.get("aliases", [])),
        _flags(entry),
        _description(entry),
    )


def bootstrap_entry_to_row(entry):
    """Convert a BOOTSTRAP_ENV_REGISTRY entry to a tuple row."""
    flags_parts = []
    if entry.get("secret"):
        flags_parts.append("secret")
    desc_parts = []
    if entry.get("secret"):
        desc_parts.append("[SECRET]")
    return (
        entry.get("key", ""),
        entry.get("type", ""),
        _default_display(entry.get("default", "")),
        "",  # min
        "",  # max
        entry.get("envKey", ""),
        "",  # configKey (bootstrap entries don't have one)
        "",  # aliases
        ", ".join(flags_parts),
        " ".join(desc_parts) if desc_parts else f"Bootstrap env ({entry.get('group', '')})",
    )


def storage_entry_to_row(entry):
    """Convert a STORAGE_SETTINGS_REGISTRY entry to a tuple row."""
    flags_parts = []
    if entry.get("secret"):
        flags_parts.append("secret")
    if entry.get("mutable"):
        flags_parts.append("mutable")
    if entry.get("computed"):
        flags_parts.append("computed")
    if entry.get("clearFlag"):
        flags_parts.append(f"clearFlag: {entry['clearFlag']}")
    allowed = entry.get("allowed", [])
    desc = ""
    if allowed:
        desc = f"Allowed: {', '.join(allowed)}"
    if entry.get("secret"):
        desc = "[SECRET] " + desc if desc else "[SECRET]"
    return (
        entry.get("key", ""),
        entry.get("type", ""),
        _default_display(entry.get("default", "")),
        "",  # min
        "",  # max
        "",  # envKey
        "",  # configKey
        "",  # aliases
        ", ".join(flags_parts),
        desc,
    )


def ui_entry_to_row(entry):
    """Convert a UI_SETTINGS_REGISTRY entry to a tuple row."""
    flags_parts = []
    if entry.get("mutable"):
        flags_parts.append("mutable")
    return (
        entry.get("key", ""),
        entry.get("type", ""),
        _default_display(entry.get("default", "")),
        "",  # min
        "",  # max
        "",  # envKey
        "",  # configKey
        "",  # aliases
        ", ".join(flags_parts),
        "",
    )


# ── Group runtime entries by uiSection ─────────────────────────────────

def group_runtime_by_section(entries):
    """Group RUNTIME entries by uiSection, returning ordered (tab_name, rows) pairs."""
    by_section = {}
    for entry in entries:
        section = entry.get("uiSection", "")
        if not section:
            section = "_uncategorized"
        by_section.setdefault(section, []).append(entry)

    result = []
    seen = set()
    for section_key in SECTION_ORDER:
        if section_key in by_section:
            tab_name = SECTION_TAB_NAMES.get(section_key, section_key.replace("-", " ").title())
            rows = [runtime_entry_to_row(e) for e in sorted(by_section[section_key], key=lambda e: e.get("key", ""))]
            result.append((tab_name, rows))
            seen.add(section_key)

    # Catch any sections not in SECTION_ORDER
    for section_key in sorted(by_section.keys()):
        if section_key not in seen:
            tab_name = section_key.replace("-", " ").title()
            rows = [runtime_entry_to_row(e) for e in sorted(by_section[section_key], key=lambda e: e.get("key", ""))]
            result.append((tab_name, rows))

    return result


# ── Group bootstrap entries by group (with subheaders) ─────────────────

def group_bootstrap_flat(entries):
    """Build the Bootstrap tab rows with subheader rows between groups."""
    by_group = {}
    for entry in entries:
        g = entry.get("group", "other")
        by_group.setdefault(g, []).append(entry)

    rows = []
    group_order = ["core", "caching", "storage", "security", "llm", "discovery", "runtime", "paths"]
    seen = set()
    for g in group_order:
        if g in by_group:
            _add_group_rows(rows, g, by_group[g])
            seen.add(g)
    for g in sorted(by_group.keys()):
        if g not in seen:
            _add_group_rows(rows, g, by_group[g])
    return rows


def _add_group_rows(rows, group_name, entries):
    """Append a subheader row then data rows for a bootstrap group."""
    # Subheader: use the key column for the group title, rest empty
    rows.append((f"--- {group_name.upper()} ---", "", "", "", "", "", "", "", "", ""))
    for entry in sorted(entries, key=lambda e: e.get("key", "")):
        rows.append(bootstrap_entry_to_row(entry))


# ── Sheet builder ──────────────────────────────────────────────────────

def style_sheet(ws, rows, tab_name):
    """Apply styling to a worksheet."""
    ws.sheet_properties.tabColor = TAB_COLORS.get(tab_name, "808080")
    ws.freeze_panes = "A2"

    # Set column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        key_val = str(row_data[0]) if row_data else ""
        is_subheader = key_val.startswith("---") and key_val.endswith("---")

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER

            if is_subheader:
                cell.font = SUBHEADER_FONT
                cell.fill = SUBHEADER_FILL
                continue

            # Conditional styling
            flags = str(row_data[8]).lower() if len(row_data) > 8 else ""
            dtype = str(row_data[1]).lower() if len(row_data) > 1 else ""
            default_val = row_data[2] if len(row_data) > 2 else ""

            if "secret" in flags:
                cell.fill = SECRET_FILL
            elif "deprecated" in flags:
                cell.fill = DEPRECATED_FILL
            elif "readonly" in flags:
                cell.fill = READONLY_FILL
            elif col_idx == 3 and dtype == "bool":
                if str(default_val).lower() == "true":
                    cell.fill = BOOL_TRUE_FILL
                elif str(default_val).lower() == "false":
                    cell.fill = BOOL_FALSE_FILL

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"


# ── Retired tab ────────────────────────────────────────────────────────

RETIRED_COLUMNS = ["Key", "Type", "Last Default", "Removed Date", "Category", "Reason"]
RETIRED_COL_WIDTHS = [38, 10, 24, 16, 20, 52]


def style_retired_sheet(ws, rows):
    """Apply styling to the Retired settings tab."""
    ws.sheet_properties.tabColor = TAB_COLORS.get("Retired", "C00000")
    ws.freeze_panes = "A2"

    for i, w in enumerate(RETIRED_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col_idx, header in enumerate(RETIRED_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER
            cell.fill = RETIRED_FILL

    ws.auto_filter.ref = f"A1:{get_column_letter(len(RETIRED_COLUMNS))}{len(rows) + 1}"


# ── Overview / Summary sheet ───────────────────────────────────────────

def add_summary_sheet(wb, all_tabs, retired_count):
    """Add a summary/overview sheet as the first tab."""
    ws = wb.create_sheet("Overview", 0)
    ws.sheet_properties.tabColor = "000000"

    title_font = Font(name="Calibri", bold=True, size=16, color="2F5496")
    subtitle_font = Font(name="Calibri", bold=True, size=12, color="404040")
    stat_font = Font(name="Calibri", size=11)
    stat_bold = Font(name="Calibri", bold=True, size=11)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 50

    row = 2
    ws.cell(row=row, column=2, value="Spec Factory -- Runtime Knobs Reference").font = title_font
    row += 1
    ws.cell(row=row, column=2, value="Auto-generated from live settingsRegistry.js (SSOT)").font = Font(name="Calibri", size=10, italic=True, color="808080")
    row += 2

    ws.cell(row=row, column=2, value="Panel / Tab").font = SUBHEADER_FONT
    ws.cell(row=row, column=2).fill = SUBHEADER_FILL
    ws.cell(row=row, column=3, value="Knob Count").font = SUBHEADER_FONT
    ws.cell(row=row, column=3).fill = SUBHEADER_FILL
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value="Registry Source").font = SUBHEADER_FONT
    ws.cell(row=row, column=4).fill = SUBHEADER_FILL
    row += 1

    total = 0
    for tab_name, tab_rows, source in all_tabs:
        # Don't count subheader rows
        count = sum(1 for r in tab_rows if not (str(r[0]).startswith("---") and str(r[0]).endswith("---")))
        total += count
        ws.cell(row=row, column=2, value=tab_name).font = stat_font
        c = ws.cell(row=row, column=3, value=count)
        c.font = stat_font
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=source).font = stat_font
        row += 1

    # Retired row
    ws.cell(row=row, column=2, value="Retired").font = stat_font
    c = ws.cell(row=row, column=3, value=retired_count)
    c.font = stat_font
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value="Manually tracked").font = stat_font
    row += 1

    row += 1
    ws.cell(row=row, column=2, value="TOTAL ACTIVE KNOBS").font = stat_bold
    c = ws.cell(row=row, column=3, value=total)
    c.font = stat_bold
    c.alignment = Alignment(horizontal="center")

    row += 2
    ws.cell(row=row, column=2, value="Legend").font = subtitle_font
    row += 1
    legends = [
        (SECRET_FILL, "Secret / API key (masked in UI)"),
        (DEPRECATED_FILL, "Deprecated (kept for backward compat)"),
        (READONLY_FILL, "Read-only (not user-editable)"),
        (BOOL_TRUE_FILL, "Boolean default = true"),
        (BOOL_FALSE_FILL, "Boolean default = false"),
        (RETIRED_FILL, "Retired (removed from registry)"),
    ]
    for fill, label in legends:
        ws.cell(row=row, column=2, value="     ").fill = fill
        ws.cell(row=row, column=3, value=label).font = stat_font
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="Source: src/shared/settingsRegistry.js (live dump via Node.js)").font = Font(name="Calibri", size=9, italic=True, color="A0A0A0")


# ── Main ───────────────────────────────────────────────────────────────

def main():
    print("Dumping live registry via Node.js...")
    data = load_registry_json()

    runtime_entries = data["runtime"]
    bootstrap_entries = data["bootstrap"]
    ui_entries = data["ui"]
    storage_entries = data["storage"]

    print(f"  RUNTIME:   {len(runtime_entries)} entries")
    print(f"  BOOTSTRAP: {len(bootstrap_entries)} entries")
    print(f"  UI:        {len(ui_entries)} entries")
    print(f"  STORAGE:   {len(storage_entries)} entries")

    # Group runtime by uiSection
    runtime_tabs = group_runtime_by_section(runtime_entries)

    # Build bootstrap rows with subheaders
    bootstrap_rows = group_bootstrap_flat(bootstrap_entries)

    # Storage and UI rows
    storage_rows = [storage_entry_to_row(e) for e in sorted(storage_entries, key=lambda e: e.get("key", ""))]
    ui_rows = [ui_entry_to_row(e) for e in sorted(ui_entries, key=lambda e: e.get("key", ""))]

    # Collect all tabs: (tab_name, rows, source_label)
    all_tabs = []
    for tab_name, rows in runtime_tabs:
        all_tabs.append((tab_name, rows, "RUNTIME_SETTINGS_REGISTRY"))
    all_tabs.append(("Bootstrap", bootstrap_rows, "BOOTSTRAP_ENV_REGISTRY"))
    all_tabs.append(("Storage", storage_rows, "STORAGE_SETTINGS_REGISTRY"))
    all_tabs.append(("UI Settings", ui_rows, "UI_SETTINGS_REGISTRY"))

    # Build workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    add_summary_sheet(wb, all_tabs, len(RETIRED_SETTINGS))

    for tab_name, rows, _source in all_tabs:
        ws = wb.create_sheet(title=tab_name)
        style_sheet(ws, rows, tab_name)

    # Retired tab
    ws = wb.create_sheet(title="Retired")
    style_retired_sheet(ws, RETIRED_SETTINGS)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT_PATH))

    active_count = sum(
        sum(1 for r in rows if not (str(r[0]).startswith("---") and str(r[0]).endswith("---")))
        for _, rows, _ in all_tabs
    )
    tab_count = len(all_tabs) + 1  # +1 for Retired
    print(f"\nWrote {OUT_PATH.relative_to(REPO_ROOT)}")
    print(f"  {active_count} active knobs across {tab_count} tabs + Overview")
    print(f"  {len(RETIRED_SETTINGS)} retired settings")


if __name__ == "__main__":
    main()
