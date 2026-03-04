from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path("implementation/ui-styling-system-standardization/review-button-color-matrix.xlsx")

HEX = {
    "blue": "#3B82F6",
    "orange": "#EA580C",
    "green": "#16A34A",
    "purple": "#9333EA",
    "violet": "#7C3AED",
    "indigo": "#4F46E5",
    "teal": "#0D9488",
    "amber": "#F59E0B",
    "red": "#DC2626",
    "sky": "#38BDF8",
    "gray": "#64748B",
}


def nhex(v: str) -> str:
    return v.strip().replace("#", "").upper()


def fit(ws) -> None:
    for i, col in enumerate(ws.columns, 1):
        w = 10
        for c in col:
            if c.value is None:
                continue
            w = max(w, min(60, len(str(c.value)) + 2))
        ws.column_dimensions[get_column_letter(i)].width = w


def style_title(ws, title: str, subtitle: str, cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True, color="1D4ED8")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, color="334155")
    ws["A2"].alignment = Alignment(wrap_text=True)


def style_head(ws, cols: int) -> None:
    fill = PatternFill("solid", fgColor="0F172A")
    font = Font(color="FFFFFF", bold=True, size=10)
    for c in range(1, cols + 1):
        cell = ws.cell(3, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def draw_grid(ws, r1: int, r2: int, cols: int) -> None:
    b = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for r in range(r1, r2 + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            cell.border = b
            if r > 3:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_table(ws, name: str, cols: int, rows: int) -> None:
    ref = f"A3:{get_column_letter(cols)}{rows}"
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(t)


def paint_swatches(ws, start_row: int, row_count: int, hex_col: int, swatch_col: int) -> None:
    for r in range(start_row, start_row + row_count):
        val = ws.cell(r, hex_col).value
        ws.cell(r, swatch_col).value = ""
        if isinstance(val, str) and val.strip().startswith("#"):
            ws.cell(r, swatch_col).fill = PatternFill("solid", fgColor=nhex(val))


def sheet(wb: Workbook, name: str, title: str, subtitle: str, headers: list[str], rows: list[tuple], table_name: str) -> None:
    ws = wb.create_sheet(name)
    cols = len(headers)
    style_title(ws, title, subtitle, cols)
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_head(ws, cols)
    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx, value)
    last = 3 + len(rows)
    add_table(ws, table_name, cols, last)
    draw_grid(ws, 3, last, cols)
    if "Current Hex" in headers and "Current Swatch" in headers:
        paint_swatches(ws, 4, len(rows), headers.index("Current Hex") + 1, headers.index("Current Swatch") + 1)
    if "Target Hex" in headers and "Target Swatch" in headers:
        paint_swatches(ws, 4, len(rows), headers.index("Target Hex") + 1, headers.index("Target Swatch") + 1)
    ws.freeze_panes = "A4"
    fit(ws)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "Review Grid + Review Components Button Color Matrix"
    ws["A1"].font = Font(size=16, bold=True, color="1D4ED8")
    ws.merge_cells("A1:F1")
    notes = [
        "Purpose: full matrix of button lanes by level (top-level, cell, cell AI, drawer) for both panels.",
        "Edit the target columns directly (Target Role / Target Hex / Target Primitive).",
        "Scope includes Review Grid, Review Components, shared CellDrawer lanes, and flags/metrics colors.",
        "Generated: 2026-03-03",
    ]
    for i, line in enumerate(notes, 3):
        ws[f"A{i}"] = line
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    headers = [
        "Lane",
        "Context",
        "Current Primitive/Class",
        "Current Hex",
        "Current Swatch",
        "Target Role (Editable)",
        "Target Hex",
        "Target Swatch",
        "Target Primitive (Editable)",
        "Notes",
    ]
    rows = [
        ("Item Run AI", "Review Grid CellDrawer", "sf-run-ai-button", HEX["purple"], "", "Run AI", HEX["purple"], "", "sf-run-ai-button", "Fixed to #9333EA"),
        ("Component/Enum Run AI", "ComponentSubTab + EnumSubTab + Panel batch", "sf-run-ai-button", HEX["purple"], "", "Run AI", HEX["purple"], "", "sf-run-ai-button", "Fixed to #9333EA"),
        ("Item Accept", "Grid CellDrawer item lane", "sf-item-accept-button", HEX["blue"], "", "Item accept primitive", HEX["blue"], "", "sf-item-accept-button", "Uses #3B82F6"),
        ("Item Confirm AI", "Grid confirm lane", "sf-confirm-button-solid", HEX["orange"], "", "Confirm primitive", HEX["orange"], "", "sf-confirm-button-solid", "Uses #EA580C"),
        ("Shared Accept", "Component/Enum shared lane", "sf-shared-accept-button", HEX["blue"], "", "Shared accept primitive", HEX["blue"], "", "sf-shared-accept-button", "Set to #3B82F6"),
        ("Shared Confirm AI", "Component/Enum shared confirm", "sf-shared-confirm-button", HEX["orange"], "", "Shared confirm primitive", HEX["orange"], "", "sf-shared-confirm-button", "Uses #EA580C"),
        ("Accepted State", "Badges/status", "green badges / sf-chip-success", HEX["green"], "", "Green (requested)", HEX["green"], "", "accepted-state token", "State only"),
        ("Flags", "Flag chips and warnings", "sf-chip-warning / sf-chip-danger", HEX["amber"], "", "Amber + Red", HEX["amber"], "", "sf-chip-warning/sf-chip-danger", "Non-button lane"),
        ("Metrics", "Meter fills", "sf-meter-fill-success/warning/danger/info", HEX["sky"], "", "Semantic ramp", HEX["sky"], "", "sf-meter-fill-*", "Already semantic"),
    ]
    sheet(
        wb,
        "Color Schematic",
        "Color Schematic",
        "High-level lane contract with editable targets.",
        headers,
        rows,
        "ColorSchematicTable",
    )

    headers = [
        "Level",
        "Area",
        "Button/Control",
        "Scope",
        "File:Line",
        "Current Primitive/Class",
        "Current Hex",
        "Current Swatch",
        "Target Role (Editable)",
        "Target Hex",
        "Target Swatch",
        "Target Primitive (Editable)",
        "Notes",
    ]
    rows = [
        ("Top", "Review toolbar", "Flagged Only", "Filter", "ReviewPage.tsx:708", "sf-chip-info sf-border-default (ON) / sf-icon-button (OFF)", HEX["blue"], "", "Filter Toggle", HEX["blue"], "", "sf-chip-info + sf-icon-button", "Uses shared toggle contract"),
        ("Top", "Review toolbar", "Approve", "Bulk", "ReviewPage.tsx:750", "sf-primary-button (pending) / sf-success-button-solid (approved+disabled)", HEX["blue"], "", "Approve lane", HEX["blue"], "", "sf-primary-button / sf-success-button-solid", "Approved state turns green and non-clickable"),
        ("Top", "Review toolbar", "Finalize", "Finalize", "ReviewPage.tsx:767", "sf-confirm-button-solid", HEX["orange"], "", "Finalize lane", HEX["orange"], "", "sf-confirm-button-solid", "Fixed 50/50 split with Approve"),
        ("Top", "Brand filter", "All", "Filter", "BrandFilterBar.tsx:24", "blue active / gray idle", HEX["blue"], "", "Filter chip", HEX["blue"], "", "filter-chip primitive", ""),
        ("Top", "Brand filter", "None", "Filter", "BrandFilterBar.tsx:34", "blue active / gray idle", HEX["blue"], "", "Filter chip", HEX["blue"], "", "filter-chip primitive", ""),
        ("Top", "Brand filter", "Brand chip", "Filter", "BrandFilterBar.tsx:49", "blue tint active / gray idle", HEX["blue"], "", "Filter chip", HEX["blue"], "", "filter-chip primitive", ""),
        ("Drawer", "DrawerShell", "Close (x)", "Utility", "DrawerShell.tsx:47", "sf-drawer-close", HEX["gray"], "", "Utility", HEX["gray"], "", "sf-drawer-close", ""),
        ("Drawer", "Manual override", "Apply", "Override", "DrawerShell.tsx:202", "sf-drawer-apply-button (soft blue -> solid hover)", HEX["blue"], "", "Override apply", HEX["blue"], "", "sf-drawer-apply-button", ""),
        ("Drawer current", "CellDrawer grid", "Confirm Item", "Item confirm", "CellDrawer.tsx:380", "sf-confirm-button-solid", HEX["orange"], "", "Confirm primitive", HEX["orange"], "", "sf-confirm-button-solid", ""),
        ("Drawer current", "CellDrawer grid", "Accept (current value)", "Item accept", "CellDrawer.tsx:400", "sf-item-accept-button", HEX["blue"], "", "Item accept primitive", HEX["blue"], "", "sf-item-accept-button", ""),
        ("Drawer AI", "CellDrawer candidates", "Run AI Review", "Item run AI", "CellDrawer.tsx:423", "sf-run-ai-button", HEX["purple"], "", "Run AI", HEX["purple"], "", "sf-run-ai-button", ""),
        ("Drawer row", "CellDrawer grid", "Accept", "Item accept", "CellDrawer.tsx:630", "sf-item-accept-button / sf-review-accepted-button", HEX["blue"], "", "Item accept primitive", HEX["blue"], "", "sf-item-accept-button", "Accepted state uses solid green"),
        ("Drawer row", "CellDrawer grid", "Confirm", "Item confirm", "CellDrawer.tsx:653", "sf-confirm-button-solid", HEX["orange"], "", "Confirm primitive", HEX["orange"], "", "sf-confirm-button-solid", ""),
        ("Drawer row", "CellDrawer shared path", "Confirm Shared", "Shared confirm", "CellDrawer.tsx:658", "sf-shared-confirm-button", HEX["orange"], "", "Shared confirm primitive", HEX["orange"], "", "sf-shared-confirm-button", ""),
        ("Drawer row", "CellDrawer shared path", "Accept Shared", "Shared accept", "CellDrawer.tsx:642", "sf-shared-accept-button / sf-review-accepted-button", HEX["blue"], "", "Shared accept primitive", HEX["blue"], "", "sf-shared-accept-button", "Accepted state uses solid green"),
    ]
    sheet(
        wb,
        "Review Grid",
        "Review Grid Button Matrix",
        "Top-level + cell + AI + drawer controls for review grid.",
        headers,
        rows,
        "ReviewGridTable",
    )

    rows = [
        ("Top", "ComponentReviewPage", "Debug LP+ID", "Utility", "ComponentReviewPage.tsx:115", "sf-icon-button / sf-chip-info", HEX["gray"], "", "Utility toggle", HEX["sky"], "", "debug-toggle primitive", ""),
        ("Top", "ComponentReviewPage", "Subtab selector", "Nav", "ComponentReviewPage.tsx:134", "sf-nav-item / sf-nav-item-active", HEX["blue"], "", "Navigation", HEX["blue"], "", "sf-nav-item", ""),
        ("Cell", "ComponentSubTab", "Linked products expand", "Cell utility", "ComponentSubTab.tsx:465", "inline blue/gray classes", HEX["blue"], "", "Cell utility", HEX["blue"], "", "linked-products primitive", ""),
        ("Cell AI", "ComponentSubTab", "Run AI", "Component run AI", "ComponentSubTab.tsx:615", "sf-run-ai-button", HEX["purple"], "", "Run AI", HEX["purple"], "", "sf-run-ai-button", ""),
        ("Panel batch", "ComponentReviewPanel", "Run AI Review All", "Component run AI", "ComponentReviewPanel.tsx:195", "sf-run-ai-button", HEX["purple"], "", "Run AI", HEX["purple"], "", "sf-run-ai-button", ""),
        ("Panel row", "ComponentReviewPanel", "Approve New", "Item accept", "ComponentReviewPanel.tsx:103", "sf-primary-button", HEX["blue"], "", "Item accept", HEX["blue"], "", "sf-primary-button", ""),
        ("Panel row", "ComponentReviewPanel", "Merge as Alias", "Secondary", "ComponentReviewPanel.tsx:111", "sf-action-button", HEX["blue"], "", "Secondary", HEX["blue"], "", "sf-action-button", ""),
        ("Panel row", "ComponentReviewPanel", "Dismiss", "Neutral", "ComponentReviewPanel.tsx:119", "sf-icon-button", HEX["gray"], "", "Neutral", HEX["gray"], "", "sf-icon-button", ""),
        ("Panel util", "ComponentReviewPanel", "Show/Hide Details", "Utility", "ComponentReviewPanel.tsx:203", "sf-icon-button", HEX["gray"], "", "Utility", HEX["gray"], "", "sf-icon-button", ""),
        ("Top", "EnumSubTab", "Field selector row", "Nav", "EnumSubTab.tsx:94", "sf-review-enum-field-*", HEX["blue"], "", "Navigation", HEX["blue"], "", "sf-review-enum-field-*", ""),
        ("Cell", "EnumSubTab", "Value row selector", "Cell select", "EnumSubTab.tsx:180", "sf-review-enum-row-*", HEX["blue"], "", "Cell select", HEX["blue"], "", "sf-review-enum-row-*", ""),
        ("Cell AI", "EnumSubTab", "AI (row)", "Enum run AI", "EnumSubTab.tsx:218", "sf-run-ai-button", HEX["purple"], "", "Run AI", HEX["purple"], "", "sf-run-ai-button", ""),
        ("Top", "EnumSubTab header", "Consistency", "LLM utility", "EnumSubTab.tsx:635", "sf-llm-soft-button", HEX["purple"], "", "LLM utility (soft)", HEX["purple"], "", "sf-llm-soft-button", ""),
        ("Top", "EnumSubTab header", "Run AI Review", "Enum run AI", "EnumSubTab.tsx:646", "sf-run-ai-button", HEX["purple"], "", "Run AI", HEX["purple"], "", "sf-run-ai-button", ""),
        ("Top", "EnumSubTab footer", "Add", "Create", "EnumSubTab.tsx:707", "sf-primary-button", HEX["blue"], "", "Create", HEX["blue"], "", "sf-primary-button", ""),
        ("Drawer", "Enum value drawer", "Remove Value", "Destructive", "EnumSubTab.tsx:742", "sf-danger-button-solid", HEX["red"], "", "Destructive", HEX["red"], "", "sf-danger-button-solid", ""),
        ("Drawer bulk", "ComponentReviewDrawer", "Accept Entire Row", "Bulk accept", "ComponentReviewDrawer.tsx:1188", "sf-primary-button", HEX["blue"], "", "Bulk accept", HEX["blue"], "", "bulk-accept primitive", ""),
        ("Drawer bulk", "ComponentReviewDrawer", "Accept All Values + Approve", "Bulk accept", "ComponentReviewDrawer.tsx:1207", "sf-drawer-apply-button (soft blue -> solid hover)", HEX["blue"], "", "Bulk approve (soft blue)", HEX["blue"], "", "sf-drawer-apply-button", ""),
        ("Drawer edit", "ComponentReviewDrawer", "Save", "Confirm", "ComponentReviewDrawer.tsx:276", "sf-confirm-button-solid", HEX["orange"], "", "Confirm primitive", HEX["orange"], "", "sf-confirm-button-solid", ""),
        ("Drawer edit", "ComponentReviewDrawer", "Apply (property)", "Override apply", "ComponentReviewDrawer.tsx:222", "sf-drawer-apply-button (soft blue -> solid hover)", HEX["blue"], "", "Override apply", HEX["blue"], "", "sf-drawer-apply-button", ""),
        ("Drawer edit", "ComponentReviewDrawer", "Save Aliases", "Confirm", "ComponentReviewDrawer.tsx:379", "sf-confirm-button-solid", HEX["orange"], "", "Confirm primitive", HEX["orange"], "", "sf-confirm-button-solid", ""),
        ("Drawer edit", "ComponentReviewDrawer", "Save Links", "Confirm", "ComponentReviewDrawer.tsx:470", "sf-confirm-button-solid", HEX["orange"], "", "Confirm primitive", HEX["orange"], "", "sf-confirm-button-solid", ""),
        ("Drawer shared", "CellDrawer in component/enum", "Accept candidate", "Shared accept", "CellDrawer.tsx:627", "sf-shared-accept-button (+green active)", HEX["blue"], "", "Shared accept primitive", HEX["blue"], "", "sf-shared-accept-button", ""),
        ("Drawer shared", "CellDrawer in component/enum", "Confirm shared", "Shared confirm", "CellDrawer.tsx:658", "sf-shared-confirm-button", HEX["orange"], "", "Shared confirm primitive", HEX["orange"], "", "sf-shared-confirm-button", ""),
    ]
    sheet(
        wb,
        "Review Components",
        "Review Components Button Matrix",
        "Component + Enum panel controls, including drawer actions.",
        headers,
        rows,
        "ReviewComponentsTable",
    )

    headers2 = [
        "Lane",
        "Grid Mapping",
        "Component/Enum Mapping",
        "Current Grid",
        "Current Grid Hex",
        "Grid Swatch",
        "Current Component",
        "Current Component Hex",
        "Component Swatch",
        "Target Role (Editable)",
        "Target Hex",
        "Target Swatch",
        "Target Primitive (Editable)",
        "Notes",
    ]
    rows2 = [
        ("Item Run AI", "CellDrawer Run AI Review", "-", "sf-run-ai-button", HEX["purple"], "", "-", "", "", "Run AI", HEX["purple"], "", "sf-run-ai-button", ""),
        ("Component/Enum Run AI", "-", "ComponentSubTab/EnumSubTab/Panel", "-", "", "", "sf-run-ai-button", HEX["purple"], "", "Run AI", HEX["purple"], "", "sf-run-ai-button", ""),
        ("Item Accept", "Accept current + Accept Item", "Approve New", "sf-item-accept-button", HEX["blue"], "", "sf-primary-button", HEX["blue"], "", "Item accept primitive", HEX["blue"], "", "sf-item-accept-button", "All accepts blue"),
        ("Item Confirm AI", "Confirm Item", "Confirm-like saves", "sf-confirm-button-solid", HEX["orange"], "", "sf-confirm-button-solid", HEX["orange"], "", "Confirm primitive", HEX["orange"], "", "sf-confirm-button-solid", ""),
        ("Shared Accept", "Shared path (not shown in grid flow)", "Shared candidate accept", "sf-shared-accept-button", HEX["blue"], "", "sf-shared-accept-button", HEX["blue"], "", "Shared accept primitive", HEX["blue"], "", "sf-shared-accept-button", ""),
        ("Shared Confirm AI", "Shared confirm path", "Shared confirm path", "sf-shared-confirm-button", HEX["orange"], "", "sf-shared-confirm-button", HEX["orange"], "", "Shared confirm primitive", HEX["orange"], "", "sf-shared-confirm-button", ""),
        ("Accepted State", "Accepted badge state", "Accepted badges/chips", "green state", HEX["green"], "", "green state", HEX["green"], "", "Accepted", HEX["green"], "", "accepted-state token", "Requested green"),
    ]
    sheet(
        wb,
        "Action Lane Matrix",
        "Action Lane Matrix",
        "4-lane action model + Run AI and Accepted state.",
        headers2,
        rows2,
        "ActionLaneTable",
    )

    headers3 = [
        "Category",
        "Surface",
        "Element",
        "Current Primitive/Class",
        "Current Hex",
        "Current Swatch",
        "Target Role (Editable)",
        "Target Hex",
        "Target Swatch",
        "Notes",
    ]
    rows3 = [
        ("Save status", "ReviewPage", "Saving/Saved/Error text", "sf-status-text-info / sf-status-text-success / sf-status-text-danger", HEX["sky"], "", "Info/Success/Error", HEX["sky"], "", "sf-status-text-*"),
        ("AI pending", "CellDrawer", "Item pending banner", "sf-review-ai-pending-banner", HEX["purple"], "", "Item AI pending", HEX["purple"], "", "sf-review-ai-pending-banner"),
        ("AI pending", "CellDrawer", "Shared pending banner", "sf-review-ai-pending-banner", HEX["purple"], "", "Shared AI pending", HEX["purple"], "", "sf-review-ai-pending-banner"),
        ("Accepted", "CellDrawer", "Accepted badge + active row", "green badge/row", HEX["green"], "", "Accepted", HEX["green"], "", "State-only"),
        ("Flags", "EnumSubTab", "AI chip + danger/warning chips", "sf-chip-accent / sf-chip-danger / sf-chip-warning", HEX["amber"], "", "AI + warning/danger", HEX["amber"], "", "semantic chips"),
        ("LLM badge", "EnumSubTab", "LLM badge next to Consistency", "Removed", HEX["gray"], "", "Removed", HEX["gray"], "", "n/a", "LLM + ? removed per current UX"),
        ("Flags", "ComponentSubTab", "Synthetic AI + flag chips", "sf-review-ai-pending-badge + sf-component-flag-chip", HEX["purple"], "", "AI + warning", HEX["amber"], "", "semantic chips"),
        ("Metrics", "ComponentReviewPanel", "Meter fills", "sf-meter-fill-success/warning/danger/info", HEX["green"], "", "Metric ramp", HEX["green"], "", "semantic"),
        ("Panel chips", "ComponentReviewPanel", "pending AI / needs review / auto-aliased", "sf-chip-accent / sf-chip-warning / sf-chip-success", HEX["blue"], "", "AI/warning/success", HEX["blue"], "", "semantic"),
    ]
    sheet(
        wb,
        "Flags & Metrics",
        "Flags and Metrics Color Matrix",
        "Non-button color lanes that still affect review panel readability.",
        headers3,
        rows3,
        "FlagsMetricsTable",
    )

    headers4 = [
        "Token / Primitive",
        "Type",
        "Light Value",
        "Light Hex",
        "Light Swatch",
        "Dark Value",
        "Dark Hex",
        "Dark Swatch",
        "Usage",
        "Source",
    ]
    rows4 = [
        ("--sf-token-accent", "Token", "#3b82f6", "#3B82F6", "", "#6366f1", "#6366F1", "", "Primary actions", "theme.css:13,94"),
        ("--sf-token-state-confirm-fg", "Token", "#ea580c", "#EA580C", "", "#fb923c", "#FB923C", "", "Shared confirm lane token", "theme.css:19,104"),
        ("--sf-token-state-run-ai-fg", "Token", "#9333ea", "#9333EA", "", "#9333ea", "#9333EA", "", "Run AI lane token", "theme.css:15,100"),
        ("--sf-token-state-item-accept-fg", "Token", "#3b82f6", "#3B82F6", "", "#3b82f6", "#3B82F6", "", "Item accept lane token", "theme.css:16,101"),
        ("--sf-token-state-shared-confirm-fg", "Token", "#ea580c", "#EA580C", "", "#ea580c", "#EA580C", "", "Shared confirm lane token", "theme.css:17,102"),
        ("--sf-token-state-shared-accept-fg", "Token", "#3b82f6", "#3B82F6", "", "#3b82f6", "#3B82F6", "", "Shared accept lane token", "theme.css:18,103"),
        ("--sf-token-state-success-fg", "Token", "#16a34a", "#16A34A", "", "#86efac", "#86EFAC", "", "Accepted state", "theme.css:21,102"),
        ("--sf-token-state-warning-fg", "Token", "#f59e0b", "#F59E0B", "", "#fcd34d", "#FCD34D", "", "Warnings", "theme.css:18,99"),
        ("--sf-token-state-error-fg", "Token", "#dc2626", "#DC2626", "", "#f87171", "#F87171", "", "Danger", "theme.css:27,108"),
        (".sf-primary-button", "Primitive", "accent background", "#3B82F6", "", "accent background", "#6366F1", "", "Primary button", "theme.css:281"),
        (".sf-confirm-button-solid", "Primitive", "confirm background", "#EA580C", "", "confirm background", "#FB923C", "", "Confirm primitive", "theme.css:337"),
        (".sf-run-ai-button", "Primitive", "run-ai background", "#9333EA", "", "run-ai background", "#9333EA", "", "Run AI primitive", "theme.css:348"),
        (".sf-item-accept-button", "Primitive", "item-accept background", "#3B82F6", "", "item-accept background", "#3B82F6", "", "Item accept primitive", "theme.css:359"),
        (".sf-shared-confirm-button", "Primitive", "shared-confirm background", "#EA580C", "", "shared-confirm background", "#EA580C", "", "Shared confirm primitive", "theme.css:370"),
        (".sf-shared-accept-button", "Primitive", "shared-accept background", "#3B82F6", "", "shared-accept background", "#3B82F6", "", "Shared accept primitive", "theme.css:381"),
        (".sf-action-button", "Primitive", "accent outline", "#3B82F6", "", "accent outline", "#6366F1", "", "Secondary action", "theme.css:270"),
        (".sf-danger-button-solid", "Primitive", "danger background", "#DC2626", "", "danger background", "#F87171", "", "Destructive button", "theme.css:303"),
        (".sf-llm-soft-button", "Primitive", "light purple tint", "#9333EA", "", "light purple tint", "#9333EA", "", "Consistency button soft style", "theme.css:392"),
        (".sf-llm-soft-badge", "Primitive", "light purple badge", "#9333EA", "", "light purple badge", "#9333EA", "", "LLM badge soft style", "theme.css:403"),
    ]
    sheet(
        wb,
        "Theme Tokens",
        "Theme Tokens and Primitives",
        "Button/status token references used by the two review panels.",
        headers4,
        rows4,
        "ThemeTokensTable",
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Workbook written: {OUT}")


if __name__ == "__main__":
    main()
